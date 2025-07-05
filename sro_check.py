import openpyxl
import os
import config
from datetime import datetime
from pars import Table_parser

months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

def sro_send(bot, mess_list, message, name):
    mess = 'Падение за месяц: \n'+'\n'.join(mess_list[0])+'\n\n'+'Уход на другие ЭТП:'+'\n'+'\n'.join(mess_list[1])
    bot.send_message(message.chat.id, mess)
    sro_path = os.getcwd()+'/report.xlsx'
    with open(sro_path,"rb") as file:
        f=file.read()
        bot.send_document(message.chat.id, f, visible_file_name=f"{name}.xlsx")



def sro_au_info(sro_name, month, year):
        current_date = datetime.strptime(f'{year}-{month}-01', "%Y-%m-%d").date()
        tb = Table_parser
        sql = f'''SELECT * FROM sro WHERE sro_name = '{sro_name}'
'''
        tb.db_connect(Table_parser)
        print(sro_name)
        cursor = tb.cursor
        cursor.execute(sql)
        sro_massiv = cursor.fetchall()
        name_set = set()
        name_etp_month_list = []
        name_etp_month_set = set()
        final_list = []
        for el in sro_massiv:
            name_set.add(el[1])
            date = el[7].replace(day=1)
            buffer_list = [el[1], el[6], el[9], date]
            name_etp_month_list.append(buffer_list)
            name_etp_month_set.add(tuple(buffer_list))

        for elem in name_etp_month_set:
            a = list(elem)
            a.append(name_etp_month_list.count(list(elem)))
            final_list.append(a)

        sorted_data = sorted(final_list, key=lambda x: x[0])

        report_list = []
        name_list = [element for element in name_set]

        for i in range(len(name_list)):
            list1 = []
            for j in range(len(final_list)):
                if name_list[i] in final_list[j]:
                    list1.append(final_list[j])

            list2 = sorted(list1, key=lambda x: x[2])    
            report_list.append(list2)


        file_path = os.getcwd()+'/report.xlsx'

        wb = openpyxl.Workbook()
        sheet = wb.active
        ws1 = wb.create_sheet('Падение')
        ws2 = wb.create_sheet('Уход')
        row_num = 2  # Начинаем со второй строки, чтобы в первую можно было врубить фильтры
        ws1_row_num = 2
        ws2_row_num = 2
        sheet.cell(row=1, column=1).value = 'ФИО АУ'
        sheet.cell(row=1, column=2).value = "ЭТП"
        sheet.cell(row=1, column=3).value = "Сумма"
        sheet.cell(row=1, column=3).value = 'Дата'  # Дата
        sheet.cell(row=1, column=4).value = 'Кол-во публикаций'
        message_list = [[], []]

        for org_data in report_list:  
            publish_all_last = []
            publish_cdt_last = []
            publish_all_now = []
            publish_cdt_now = []
            date_cdt_now = []
            date_cdt_last = []
            date_all_now = []
            etp_none_cdt_now = []
            etp_none_cdt_last = []
            etp_cdt_now = []
            etp_cdt_last = []
            sum_cdt_last = []
            sum_all_last = []
            sum_cdt_now = []
            sum_all_now = []
            for record in org_data:   
                sheet.cell(row=row_num, column=1).value = record[0]  # ФИО АУ
                sheet.cell(row=row_num, column=2).value = record[1]  # ETP
                sheet.cell(row=row_num, column=3).value = record[2] #сумма
                sheet.cell(row=row_num, column=3).value = record[3].strftime('%Y-%m-%d')  # Дата
                sheet.cell(row=row_num, column=4).value = record[4]  # Количество публикаций

                if 'ентр дистанционных торгов' in str(record[1]) and record[3].month == current_date.month and current_date.year == record[3].year:
                    publish_cdt_now.append(record[4])
                    date_cdt_now.append(str(record[3]))
                    sum_cdt_now.append(record[2])
                    etp_cdt_now.append(record[1])

                elif 'ентр дистанционных торгов' in str(record[1]) and current_date.month - record[3].month == 1 or 'ентр дистанционных торгов' in str(record[1]) and record[3].month == 12 and current_date.month == 1 and current_date.year - record[3].year == 1:

                    publish_cdt_last.append(record[4])
                    date_cdt_last.append(str(record[3]))
                    sum_cdt_last.append(record[2])
                    etp_cdt_last.append(record[1])

                elif current_date.month == record[3].month and current_date.year == record[3].year:
                    etp_none_cdt_now.append(record[1])
                    publish_all_now.append(record[4])
                    date_all_now.append(str(record[3]))
                    sum_all_now.append(record[2])
                    
                elif current_date.month - record[3].month == 1 or record[3].month == 12 and current_date.month == 1 and current_date.year - record[3].year == 1:
                    publish_all_last.append(record[4])
                    etp_none_cdt_last.append(record[1])
                    
                row_num += 1



            if len(etp_none_cdt_now)>0 and len(etp_cdt_last) > 0 and len(etp_cdt_now) > 0:
                sum1 = sum(publish_cdt_last)+len(etp_none_cdt_last)
                sum2 = sum(publish_cdt_now)+len(etp_none_cdt_now)
                ratio_last = sum(publish_cdt_last) / sum1 if sum(publish_all_last) != 0 else 0
                ratio_now = sum(publish_cdt_now) / sum2 if sum(publish_all_now) != 0 else 0
                if ratio_last > ratio_now:
                    a = f'Прошлый месяц - ЦДТ - {sum(publish_cdt_last)}, всего - {sum1}, процент ЦДТ - {str(ratio_last*100)[0:5]}%'
                    b = f'Отчетный месяц - ЦДТ - {sum(publish_cdt_now)}, всего - {sum2}, процент ЦДТ - {str(ratio_now*100)[0:5]}%'
                    ws1.cell(row=1, column=1).value = 'ФИО АУ'
                    ws1.cell(row=1, column=2).value = 'Прошлый месяц'
                    ws1.cell(row=1, column=3).value = 'Отчётный месяц'
                    ws1.cell(row=ws1_row_num, column=1).value = org_data[0][0]
                    ws1.cell(row=ws1_row_num, column=2).value = a
                    ws1.cell(row=ws1_row_num, column=3).value = b
                    message_list[0].append(org_data[0][0])
                    message_list[0].append(a)
                    message_list[0].append(b)
                    
                    ws1_row_num += 1


            if len(date_cdt_last) > 0 and len(date_cdt_now) == 0 and len(date_all_now) > 0:
                ws2.cell(row=1, column=1).value = 'ФИО АУ'
                ws2.cell(row=1, column=2).value = 'ЭТП, куда перешел в этом месяце'
                ws2.cell(row=ws2_row_num, column=1).value = org_data[0][0]
                ws2.cell(row=ws2_row_num, column=2).value = ', '.join(etp_none_cdt_now)
                ws2_row_num += 1
                message_list[1].append(org_data[0][0]+':')
                etp_set = set()
                for i in range(len(etp_none_cdt_now)):
                    a = (etp_none_cdt_now[i], etp_none_cdt_now.count(etp_none_cdt_now[i]))
                    etp_set.add(a)
                etp_list = [list(el) for el in etp_set]

                for j in range(len(etp_list)):
                    message_list[1].append(f'{etp_list[j][0]}: {etp_list[j][1]}')

                message_list[1].append('\n')

        wb.save(file_path)
        return message_list


def sro_without_accr(month, year):
    current_date = datetime.strptime(f'{year}-{month}-01', "%Y-%m-%d").date()
    tb = Table_parser
    placeholders = ','.join(['%s'] * len(config.sro_names))
    sql = f'''SELECT * FROM sro WHERE sro_name NOT IN ({placeholders})'''

    tb.db_connect(Table_parser)
        
    cursor = tb.cursor
    cursor.execute(sql, tuple(config.sro_names))
    sro_massiv = cursor.fetchall()
    
    final_list = []

    for el in sro_massiv:
        date = el[7].replace(day=1)
        if date == current_date:
            buffer_list = [el[5], el[1], el[6]] #СРО, ФИО/наименование, ЭТП
            final_list.append(buffer_list)
    
    f_list = []

    for el in final_list:
        if el[0] in ['', ' ', 'не состоит в сро']:
            f_list.append((el[1], el[2]))
        else:
            f_list.append((el[0], el[2]))

    f_list2 = []

    for el in f_list:
        f_list2.append((el, f_list.count(el)))
      
    f_set = set(f_list2)
    f_list = list(f_set)
    list2 = sorted(f_list, key=lambda x: x[0][0])  
    
    sro_dict = {}

    # Группируем данные по СРО
    for (sro, etp), count in list2:
        if sro not in sro_dict:
            sro_dict[sro] = []
        sro_dict[sro].append((etp, count))

    # Формируем вывод в нужном формате
    result = []
    for sro in sro_dict:  # или sorted(sro_dict) для сортировки по алфавиту
        result.append(sro)
        for etp, count in sro_dict[sro]:
            result.append(f"{etp}: {count}")
        result.append("")  # Пустая строка между СРО

    res = "\n".join(result[:-1])
    return res